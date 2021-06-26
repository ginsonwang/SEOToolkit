import time
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from .conf import webhook_key

# TODO 爱站地址直接访问会超时，尝试通过点击路径方式到达目标页面

def aizhan_monitor(mode='run'):
    wb = load_workbook('爱站数据监控.xlsx')
    wb.save('爱站数据监控_bak.xlsx')

    with sync_playwright() as p:
        if mode == 'run':
            chrome = p.chromium.launch()
        else:
            mode == 'test'
            chrome = p.chromium.launch(headless=False)
        page = chrome.new_page()

        print('# 爱站关键词总量监控')
        # page.goto('https://baidurank.aizhan.com/baidu/jiwu.com/')
        page.goto('https://www.aizhan.com')
        time.sleep(1)
        page.click('input[id="domain"]')
        time.sleep(0.5)
        page.fill('input[id="domain"]', 'jiwu.com')
        time.sleep(1)
        page.click('input[value="查询"]')
        with page.expect_popup() as popup_info:
            page.click('a[id="cc1"]')
        page1 = popup_info.value

        sheet = wb['吉屋网总词量']
        row = sheet.max_row + 1
        line = [time.strftime('%Y/%m/%d')]
        line.append(page1.inner_text('#cc1').replace(',', ''))
        for i in range(0, 10):
            line.append(page1.inner_text('#f_%s' % i))
        line.insert(7, page1.inner_text('#cc2').replace(',', ''))
        line.insert(1, str(int(line[1]) + int(line[7])))
        for i, v in enumerate(line):
            c = sheet.cell(row=row, column=i+1)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.value = v
        # wb.save('爱站数据监控.xlsx')
        print('# 数据已采集\n')


        print('# 吉屋网子目录爱站关键词量监控')
        channels = {
            'pc': [
                'zhuanti', 'loupan', 'news', 'fangjia', 'esf', 'wenda', 'info', 'detail', 'huxing', 'jiaotong', 'tu', 'gongjijin', 'guwen', 'lpjd', 'xqfangyuan', 'jjr', 'calculator'
            ],
            'h5': [
                'zhuanti', 'wenda', 'www', 'calculator'
            ]
        }
        az_jiwu_chl_word = [time.strftime('%Y/%m/%d')]
        for platform in channels.keys():
            for chl in channels[platform]:
                if platform == 'pc':
                    url = 'https://baidurank.aizhan.com/baidu/jiwu.com/%s/' % chl
                else:
                    url = 'https://baidurank.aizhan.com/mobile/jiwu.com/%s/' % chl
                page.goto(url)
                az_jiwu_chl_word.append(
                    page.inner_text(".baidurank-back .red", timeout=3000)
                )
                print(
                    platform + '\t' + chl + '\t' + 
                    page.inner_text(".baidurank-back .red", timeout=3000)
                )
                time.sleep(int(time.strftime('%S'))%2 + 3)
        
        sheet = wb['吉屋网各频道词量']
        row = sheet.max_row + 1
        for i, v in enumerate(az_jiwu_chl_word):
            c = sheet.cell(row=row, column=i+1)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.value = v
        # wb.save('爱站数据监控.xlsx')
        print('# 数据已采集\n')


        print('# 吉屋网及竞品百度权重、词量监控')
        sheet = wb['竞品权重']
        brs = [time.strftime('%Y/%m/%d')]
        h5_brs = []
        words = [time.strftime('%Y/%m/%d')]
        h5_words = []
        for site in [x.value for x in sheet['B3': 'X3'][0]]:
            page.goto('https://baidurank.aizhan.com/baidu/%s/' % site)
            pc_br = page.get_attribute('.top.active .ip ul li img', 'src')
            pc_br = pc_br.replace('//statics.aizhan.com/images/br/', '').replace('.png', '')
            brs.append(pc_br)
            h5_br = page.get_attribute('.top.m .ip ul li img', 'src')
            h5_br = h5_br.replace('//statics.aizhan.com/images/mbr/', '').replace('.png', '')
            h5_brs.append(h5_br)

            pc_word = page.inner_text('#cc1').replace(',', '')
            words.append(pc_word)
            h5_word = page.inner_text('#cc2').replace(',', '')
            h5_words.append(h5_word)

            time.sleep(int(time.strftime('%S'))%2 + 3)

        chrome.close()

        brs += h5_brs
        words += h5_words

        row = sheet.max_row + 1

        for i, v in enumerate(brs):
            c = sheet.cell(row=row, column=i+1)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.value = v
        
        sheet = wb['竞品词量']
        for i, v in enumerate(words):
            c = sheet.cell(row=row, column=i+1)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.value = v
        
        wb.save('爱站数据监控.xlsx')
        print('# 数据已采集\n')
    
    wb.close()

    # 发送群消息提醒
    import requests
    import json

    webhook = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=%s' % webhook_key

    message = {
        "msgtype": "text",
        "text": {
            "content": "%s 吉屋及竞品爱站权重、关键词数据已更新。"
            % (time.strftime('%Y/%m/%d'))
        }
    } 

    requests.post(webhook, data=json.dumps(message))

    # 发送文件
    headers = {
        'Content-Type': 'multipart/form-data',
    }
    params = (
        ('key', webhook_key),
        ('type', 'file')
    )
    fb = open('爱站数据监控.xlsx', 'rb')
    rsp = requests.post(
        'https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key=%s&type=file' % webhook_key,
        headers=headers, params=params, files={'file': fb}
    )
    media_id = rsp.json()['media_id']

    message = {
        "msgtype": "file",
        "file": {
            "media_id": media_id
        }
    } 
    requests.post(webhook, data=json.dumps(message))